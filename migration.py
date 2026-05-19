"""
================================================================================
MIGRATION xlsx -> sqlite : Tableau de bord Futsal
================================================================================

Lit ton fichier Tableau_de_bord_EDF-U19.xlsx et crée une base SQLite propre
selon le schéma défini dans schema.sql.

USAGE :
    python migration.py

PRÉREQUIS :
    pip install openpyxl

Le script :
  1. Crée la base SQLite (vide) à partir de schema.sql
  2. Crée l'équipe "EDF U19 Futsal"
  3. Importe les joueurs depuis la feuille JOUEURS
  4. Importe les matchs depuis MATCHS_REF
  5. Importe les performances depuis BASE_DATA (stats brutes uniquement)
  6. Importe les notes manuelles depuis EVAL_MATCH
  7. Importe la grille de coefficients depuis CALCUL_NOTE
  8. Vérifie la cohérence à la fin
================================================================================
"""

import openpyxl
import sqlite3
import os
import sys
from pathlib import Path

# --- CONFIG ---
XLSX_PATH = "Tableau_de_bord_EDF-U19.xlsx"
DB_PATH = "futsal.db"
SCHEMA_PATH = "schema.sql"
EQUIPE_NOM = "EDF U19 Futsal"
EQUIPE_CATEGORIE = "U19"
EQUIPE_SAISON = "2025-2026"


def supprimer_base_existante():
    """Repart de zéro pour pouvoir relancer le script sans erreur."""
    if os.path.exists(DB_PATH):
        os.remove(DB_PATH)
        print(f"[reset] Ancienne base {DB_PATH} supprimée")


def creer_schema(conn):
    """Exécute schema.sql pour créer toutes les tables et vues."""
    with open(SCHEMA_PATH, "r", encoding="utf-8") as f:
        conn.executescript(f.read())
    print("[schema] Tables et vues créées")


def creer_equipe(conn):
    """Crée l'équipe et retourne son ID."""
    cur = conn.execute(
        "INSERT INTO equipe (nom, categorie, saison) VALUES (?, ?, ?)",
        (EQUIPE_NOM, EQUIPE_CATEGORIE, EQUIPE_SAISON)
    )
    eid = cur.lastrowid
    print(f"[equipe] '{EQUIPE_NOM}' créée (id={eid})")
    return eid


def importer_joueurs(conn, ws, equipe_id):
    """Importe la feuille JOUEURS. Header attendu :
    Joueur | Numero | Poste | Pied | Club | Taille_cm | Poids_kg
    """
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    # On suppose l'ordre fixe, mais on log si surprise
    expected = ["Joueur", "Numero", "Poste", "Pied", "Club", "Taille_cm", "Poids_kg"]
    if list(header[:7]) != expected:
        print(f"[warn] Header JOUEURS inattendu : {header}")

    count = 0
    joueurs_map = {}  # nom -> joueur_id
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        nom, numero, poste, pied, club, taille, poids = (list(row) + [None] * 7)[:7]
        cur = conn.execute(
            """INSERT INTO joueur (equipe_id, nom, numero, poste, pied, club, taille_cm, poids_kg)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
            (equipe_id, nom.strip(), numero, poste, pied, club, taille, poids)
        )
        joueurs_map[nom.strip()] = cur.lastrowid
        count += 1
    print(f"[joueurs] {count} joueurs importés")
    return joueurs_map


def importer_matchs(conn, ws, equipe_id):
    """Importe la feuille MATCHS_REF. Header attendu :
    Match_ID | Adversaire | Match_No | Date_match | Competition | Lieu | Score_France | Score_Adverse
    """
    rows = list(ws.iter_rows(values_only=True))
    count = 0
    matchs_map = {}  # code -> match_id
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        code, adv, no, date, compet, lieu, sp, sc = (list(row) + [None] * 8)[:8]
        cur = conn.execute(
            """INSERT INTO match (equipe_id, code, adversaire, match_no, date_match,
                                  competition, lieu, score_pour, score_contre)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (equipe_id, code.strip(), adv, no, date, compet, lieu,
             int(sp or 0), int(sc or 0))
        )
        matchs_map[code.strip()] = cur.lastrowid
        count += 1
    print(f"[matchs] {count} matchs importés")
    return matchs_map


def _int(v):
    """Convertit en int, None si vide. Tolère les flottants type 5.0."""
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


def _float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def importer_performances(conn, ws, joueurs_map, matchs_map):
    """Importe la feuille BASE_DATA. On ne prend QUE les stats brutes :
    les colonnes _40, _pct, relances_total etc. sont recalculées par la vue.
    """
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    idx = {col: i for i, col in enumerate(header)}

    # Mapping colonne_xlsx -> colonne_sql pour les stats brutes uniquement
    mapping = {
        "Buts": "buts",
        "Passes_decisives": "passes_decisives",
        "Tirs_total": "tirs_total",
        "Tirs_cadres": "tirs_cadres",
        "Tirs_hors_cadre": "tirs_hors_cadre",
        "Tirs_contres": "tirs_contres",
        "Poteau_barre": "poteau_barre",
        "Pertes_de_balles": "pertes_de_balles",
        "Passes_loupees": "passes_loupees",
        "Ballons_rendus": "ballons_rendus",
        "Interceptions_adv": "interceptions_adv",
        "Duels_perdus": "duels_perdus",
        "Erreurs_techniques": "erreurs_techniques",
        "Recuperations_adv": "recuperations_adv",
        "Duels_OFF_gagnes": "duels_off_gagnes",
        "Duels_OFF_tentes": "duels_off_tentes",
        "Duels_DEF_gagnes": "duels_def_gagnes",
        "Duels_DEF_tentes": "duels_def_tentes",
        "Fautes_commises": "fautes_commises",
        "Fautes_subies": "fautes_subies",
        "Interceptions": "interceptions",
        "Recuperations": "recuperations",
        "Buts_encaisses": "buts_encaisses",
        "Arrets": "arrets",
        "Relances_faciles_reussies": "relances_faciles_reussies",
        "Relances_faciles_loupees": "relances_faciles_loupees",
        "Relances_difficiles_reussies": "relances_difficiles_reussies",
        "Relances_difficiles_loupees": "relances_difficiles_loupees",
    }

    count = 0
    ignores = 0
    for row in rows[1:]:
        if not row or not row[idx.get("Match_ID", 0)]:
            continue
        match_code = str(row[idx["Match_ID"]]).strip()
        joueur_nom = str(row[idx["Joueur"]]).strip()

        match_id = matchs_map.get(match_code)
        joueur_id = joueurs_map.get(joueur_nom)

        if not match_id:
            print(f"[warn] Match introuvable : '{match_code}' — ligne ignorée")
            ignores += 1
            continue
        if not joueur_id:
            print(f"[warn] Joueur introuvable : '{joueur_nom}' — ligne ignorée")
            ignores += 1
            continue

        # Champs de contexte
        numero_match = _int(row[idx["Numero"]])
        poste_match = row[idx["Poste"]]
        role = row[idx.get("Role", -1)] if idx.get("Role", -1) >= 0 else "Joueur"
        temps_min = _float(row[idx["Temps_jeu_min"]])
        temps_txt = row[idx["Temps_jeu_txt"]]

        # Stats brutes
        stats = {}
        for col_xlsx, col_sql in mapping.items():
            if col_xlsx in idx:
                stats[col_sql] = _int(row[idx[col_xlsx]])

        # Construction dynamique de l'INSERT
        cols = ["match_id", "joueur_id", "numero_match", "poste_match", "role",
                "temps_jeu_min", "temps_jeu_txt"] + list(stats.keys())
        vals = [match_id, joueur_id, numero_match, poste_match, role,
                temps_min, temps_txt] + list(stats.values())

        placeholders = ",".join(["?"] * len(cols))
        try:
            conn.execute(
                f"INSERT INTO performance ({','.join(cols)}) VALUES ({placeholders})",
                vals
            )
            count += 1
        except sqlite3.IntegrityError as e:
            print(f"[warn] Doublon {match_code} / {joueur_nom} : {e}")
            ignores += 1

    print(f"[perf] {count} performances importées ({ignores} ignorées)")


def importer_notes(conn, ws, joueurs_map, matchs_map):
    """Importe la feuille EVAL_MATCH. Format attendu (à partir de la ligne 4) :
    Match_ID | Libellé | Joueur | Note | Bonus
    """
    rows = list(ws.iter_rows(values_only=True))
    # Trouver la ligne header "Match_ID"
    start = None
    for i, row in enumerate(rows):
        if row and row[0] == "Match_ID":
            start = i + 1
            break
    if start is None:
        print("[notes] Aucun header trouvé, import ignoré")
        return

    count = 0
    for row in rows[start:]:
        if not row or not row[0]:
            continue
        code = str(row[0]).strip()
        # Skip les lignes d'aide en fin de feuille
        if not code.startswith(("FRA_", "MATCH_", "EDF_")) and "_M" not in code:
            continue
        joueur_nom = str(row[2]).strip() if row[2] else None
        note = _int(row[3])

        if not joueur_nom or note is None:
            continue
        if note not in (0, 5, 10):
            continue

        match_id = matchs_map.get(code)
        joueur_id = joueurs_map.get(joueur_nom)
        if not match_id or not joueur_id:
            continue

        try:
            conn.execute(
                """INSERT INTO note_match (match_id, joueur_id, note)
                   VALUES (?, ?, ?)""",
                (match_id, joueur_id, note)
            )
            count += 1
        except sqlite3.IntegrityError:
            pass  # doublon : ignoré silencieusement
    print(f"[notes] {count} notes importées")


def importer_coefficients(conn):
    """Insère la grille de coefficients depuis CALCUL_NOTE (en dur ici car
    c'est plus simple et plus sûr que de parser la feuille mise en forme)."""
    parametres = [
        ("note_depart",     10.0, "Note de départ avant ajustements"),
        ("bonus_victoire",   3.0, "Bonus appliqué si victoire"),
        ("bonus_nul",        1.5, "Bonus appliqué si match nul"),
        ("bonus_defaite",    0.0, "Bonus appliqué si défaite"),
    ]
    for cle, val, desc in parametres:
        conn.execute("INSERT INTO parametre (cle, valeur, description) VALUES (?, ?, ?)",
                     (cle, val, desc))

    coefficients = [
        # Offensif
        ("buts",              1.5,  "OFFENSIF", "Buts"),
        ("passes_decisives",  1.0,  "OFFENSIF", "Passes décisives"),
        ("tirs_cadres",       0.75, "OFFENSIF", "Tirs cadrés"),
        ("tirs_contres",      0.25, "OFFENSIF", "Tirs contrés"),
        ("tirs_hors_cadre",   0.1,  "OFFENSIF", "Tirs hors cadre"),
        ("duels_off_gagnes",  0.5,  "OFFENSIF", "Duels OFF gagnés"),
        ("fautes_subies",     0.5,  "OFFENSIF", "Fautes subies"),
        # Défensif
        ("interceptions",     0.5,  "DEFENSIF", "Interceptions"),
        ("recuperations",     0.5,  "DEFENSIF", "Récupérations"),
        ("duels_def_gagnes",  0.5,  "DEFENSIF", "Duels DEF gagnés"),
        # Négatif
        ("interceptions_adv", -0.5, "NEGATIF",  "Interceptions adverses subies"),
        ("recuperations_adv", -0.5, "NEGATIF",  "Récupérations adverses subies"),
        ("duels_perdus",      -0.5, "NEGATIF",  "Duels perdus"),
        ("fautes_commises",   -0.5, "NEGATIF",  "Fautes commises"),
        ("passes_loupees",    -0.25,"NEGATIF",  "Passes loupées"),
        ("ballons_rendus",    -0.25,"NEGATIF",  "Ballons rendus"),
        ("erreurs_techniques",-0.25,"NEGATIF",  "Erreurs techniques"),
    ]
    for action, coef, famille, libelle in coefficients:
        conn.execute(
            "INSERT INTO coefficient (action, coef, famille, libelle) VALUES (?, ?, ?, ?)",
            (action, coef, famille, libelle)
        )
    print(f"[coef] {len(parametres)} paramètres + {len(coefficients)} coefficients importés")


def verifier(conn):
    """Quelques requêtes de contrôle pour valider l'import."""
    print("\n" + "=" * 60)
    print("VÉRIFICATION")
    print("=" * 60)

    cur = conn.execute("SELECT COUNT(*) FROM joueur")
    print(f"Joueurs en base       : {cur.fetchone()[0]}")
    cur = conn.execute("SELECT COUNT(*) FROM match")
    print(f"Matchs en base        : {cur.fetchone()[0]}")
    cur = conn.execute("SELECT COUNT(*) FROM performance")
    print(f"Performances en base  : {cur.fetchone()[0]}")
    cur = conn.execute("SELECT COUNT(*) FROM note_match")
    print(f"Notes manuelles       : {cur.fetchone()[0]}")

    print("\nBilan équipe (vue v_equipe_bilan) :")
    cur = conn.execute("SELECT equipe, matchs, victoires, nuls, defaites, buts_pour, buts_contre FROM v_equipe_bilan")
    for row in cur.fetchall():
        print(f"  {row[0]} : {row[1]} matchs, {row[2]}V {row[3]}N {row[4]}D, {row[5]} buts pour / {row[6]} contre")

    print("\nTop 5 buteurs (vue v_performance) :")
    cur = conn.execute("""
        SELECT joueur, SUM(buts) AS total_buts
        FROM v_performance
        GROUP BY joueur
        HAVING total_buts > 0
        ORDER BY total_buts DESC
        LIMIT 5
    """)
    for row in cur.fetchall():
        print(f"  {row[0]:20s} : {row[1]} but(s)")

    print("\nExemple per 40 min (top buts/40) :")
    cur = conn.execute("""
        SELECT joueur, match_code, ROUND(buts_40, 2) AS b40, temps_jeu_min
        FROM v_performance
        WHERE buts_40 IS NOT NULL AND buts > 0
        ORDER BY buts_40 DESC
        LIMIT 5
    """)
    for row in cur.fetchall():
        print(f"  {row[0]:20s} ({row[1]:18s}) : {row[2]} buts/40 sur {row[3]} min")


def main():
    if not os.path.exists(SCHEMA_PATH):
        print(f"ERREUR : {SCHEMA_PATH} introuvable. Place-le dans le même dossier.")
        sys.exit(1)
    if not os.path.exists(XLSX_PATH):
        print(f"ERREUR : {XLSX_PATH} introuvable. Place-le dans le même dossier.")
        sys.exit(1)

    supprimer_base_existante()
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")

    try:
        creer_schema(conn)
        equipe_id = creer_equipe(conn)

        wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
        joueurs_map = importer_joueurs(conn, wb["JOUEURS"], equipe_id)
        matchs_map = importer_matchs(conn, wb["MATCHS_REF"], equipe_id)
        importer_performances(conn, wb["BASE_DATA"], joueurs_map, matchs_map)
        importer_notes(conn, wb["EVAL_MATCH"], joueurs_map, matchs_map)
        importer_coefficients(conn)

        conn.commit()
        verifier(conn)
        print(f"\n✓ Base créée : {Path(DB_PATH).resolve()}")
    finally:
        conn.close()


if __name__ == "__main__":
    main()
